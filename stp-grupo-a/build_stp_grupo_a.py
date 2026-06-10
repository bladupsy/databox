#!/usr/bin/env python3
"""Fetch official STP Grupo A indicators and build HTML dashboard."""

import json
import math
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

API = "https://apis.datos.gob.ar/series/api/series/"
SEARCH = "https://apis.datos.gob.ar/series/api/search"
ROOT = Path(__file__).parent
OUT = ROOT / "stp-grupo-a.html"
JSON_OUT = ROOT / "data" / "stp-grupo-a-data.json"
DATA_DIR = ROOT / "data" / "deuda"

START = "2023-01-01"
END = "2026-12-31"

# Create a robust HTTP session with retries and backoff
SESSION = requests.Session()
retry_strategy = Retry(
    total=5,
    backoff_factor=1,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["GET"]
)
adapter = HTTPAdapter(max_retries=retry_strategy)
SESSION.mount("http://", adapter)
SESSION.mount("https://", adapter)


def fetch_series(series_id, collapse=None, limit=5000):
    params = {"ids": series_id, "start_date": START, "end_date": END, "limit": limit}
    if collapse:
        params["collapse"] = collapse
    r = SESSION.get(API, params=params, timeout=90)
    r.raise_for_status()
    data = r.json()
    pts = [(p[0][:7], p[1]) for p in data.get("data", []) if p[1] is not None]
    meta = series_meta(series_id)
    return pts, meta


def series_meta(series_id):
    r = SESSION.get(SEARCH, params={"q": series_id, "limit": 1}, timeout=30)
    items = r.json().get("data", [])
    if not items:
        return {"id": series_id, "description": series_id, "source": "—", "units": "—", "dataset": "—"}
    f = items[0]["field"]
    ds = items[0].get("dataset", {})
    return {
        "id": f["id"],
        "description": f["description"],
        "source": ds.get("source", "—"),
        "units": f.get("units", "—"),
        "dataset": ds.get("title", "—"),
        "frequency": f.get("frequency", "—"),
    }


def mom_pct(points):
    out = []
    prev = None
    for date, val in points:
        if prev is not None and prev != 0:
            out.append((date, (val / prev - 1) * 100))
        prev = val
    return out


def forward_fill_monthly(points, value_key="value"):
    """Forward-fill irregular points to monthly YYYY-MM."""
    if not points:
        return []
    by_month = {d: v for d, v in points}
    months = sorted(by_month.keys())
    start = months[0]
    end = months[-1]
    y, m = map(int, start.split("-"))
    ey, em = map(int, end.split("-"))
    out = []
    last = None
    while (y, m) <= (ey, em):
        key = f"{y:04d}-{m:02d}"
        if key in by_month:
            last = by_month[key]
        if last is not None:
            out.append((key, last))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return out


def expand_quarterly_to_monthly(quarterly_points):
    out = []
    for date, val in quarterly_points:
        y, m = map(int, date.split("-")[:2])
        # assign quarter end months
        qm = {1: [1, 2, 3], 4: [4, 5, 6], 7: [7, 8, 9], 10: [10, 11, 12]}[m]
        for mm in qm:
            out.append((f"{y:04d}-{mm:02d}", val))
    return sorted(out, key=lambda x: x[0])


def ratio_series(num_pts, den_pts):
    den = dict(den_pts)
    out = []
    for date, num in num_pts:
        if date in den and den[date]:
            out.append((date, num / den[date] * 100))
    return out


def read_deuda_pib_annual():
    """Deuda bruta administración central / PIB (%) — Ministerio de Economía."""
    files = [
        DATA_DIR / "deuda_publica_31-12-2023.xlsx",
        DATA_DIR / "deuda_publica_31-12-2024.xlsx",
        DATA_DIR / "deuda_publica_31-12-2025.xlsx",
    ]
    annual = {}
    for fp in files:
        if not fp.exists():
            continue
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        ws = wb["A.4.7"]
        headers = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
        values = list(ws.iter_rows(min_row=10, max_row=10, values_only=True))[0]
        for h, v in zip(headers, values):
            if h is None or v is None:
                continue
            m = re.match(r"(\d{4})", str(h).strip())
            if m and isinstance(v, (int, float)):
                annual[int(m.group(1))] = float(v) * 100
    monthly = []
    years = sorted(annual.keys())
    max_year = max(years) if years else 2025
    for year in range(min(years), max(max_year, 2026) + 1):
        pct = annual.get(year, annual.get(max_year))
        if pct is None:
            continue
        for month in range(1, 13):
            monthly.append((f"{year:04d}-{month:02d}", pct))
    return monthly, {
        "id": "MECON_A.4.7",
        "description": "Deuda bruta de la Administración Central en relación al PIB",
        "source": "Secretaría de Finanzas, Ministerio de Economía",
        "units": "Porcentaje del PIB",
        "dataset": "Informes trimestrales de la deuda pública (hoja A.4.7)",
        "frequency": "Anual (replicada mensualmente)",
        "note": "2026 utiliza el último dato anual publicado (2025) hasta que MECON publique el cierre 2026.",
        "url": "https://www.argentina.gob.ar/economia/finanzas/deudapublica/informes-trimestrales-de-la-deuda",
    }


def build_dataset():
    indicators = []

    # 1. IPC mensual — INDEC
    ipc, ipc_meta = fetch_series("145.3_INGNACUAL_DICI_M_38")
    indicators.append({
        "key": "ipc",
        "title": "Tasa de inflación mensual (IPC nivel general)",
        "etapas": [1, 3],
        "points": [(d, v * 100) for d, v in ipc],
        "meta": ipc_meta,
        "chart_type": "line",
        "y_label": "% mensual",
        "intro": "El IPC nacional mide la variación mensual de precios al consumidor publicada por el INDEC.",
    })

    # 2. Resultado fiscal primario — MECON
    rf_p, rf_p_meta = fetch_series("452.3_RESULTADO_RIO_0_M_18_54")
    indicators.append({
        "key": "rf_primario",
        "title": "Resultado fiscal primario",
        "etapas": [1, 2, 3],
        "points": rf_p,
        "meta": rf_p_meta,
        "chart_type": "bar",
        "y_label": "Millones de $ corrientes",
        "intro": "Resultado primario del sector público nacional (ingresos menos gastos, sin intereses). Fuente: IMIG, Ministerio de Economía.",
    })

    # 3. Resultado fiscal global — MECON
    rf_g, rf_g_meta = fetch_series("452.3_RESULTADO_ERO_0_M_20_25")
    indicators.append({
        "key": "rf_global",
        "title": "Resultado fiscal global (resultado económico)",
        "etapas": [1, 2, 3],
        "points": rf_g,
        "meta": rf_g_meta,
        "chart_type": "bar",
        "y_label": "Millones de $ corrientes",
        "intro": "Resultado económico mensual del sector público nacional (incluye intereses). Fuente: IMIG.",
    })

    # 4. Bienes personales — MECON (recaudación mensual; alícuota legal es discontinua)
    bp, bp_meta = fetch_series("452.2_BIENES_PERLES_0_T_17_26")
    indicators.append({
        "key": "bienes_personales",
        "title": "Impuesto a los Bienes Personales — recaudación mensual",
        "etapas": [1, 3],
        "points": bp,
        "meta": {**bp_meta, "note": "Serie de recaudación (millones de $). La alícuota legal varía por tramo; no existe serie mensual oficial de tasa impositiva única."},
        "chart_type": "bar",
        "y_label": "Millones de $",
        "intro": "Recaudación mensual del Impuesto a los Bienes Personales (IMIG). Para alícuotas vigentes ver normativa AFIP/MECON.",
    })

    # 5. Ingresos tributarios — MECON/AFIP
    trib_total, trib_total_meta = fetch_series("142.3_TOTAL_2001_M_26")
    iva, iva_meta = fetch_series("142.3_IVA_2001_M_3")
    gan, gan_meta = fetch_series("142.3_GANAN_2001_M_9")
    pib_q, _ = fetch_series("9.2_PPC_2004_T_22")
    pib_m = expand_quarterly_to_monthly(pib_q)
    presion = ratio_series(trib_total, pib_m)

    indicators.append({
        "key": "ingresos_tributarios_total",
        "title": "Ingresos tributarios totales",
        "etapas": [1, 3],
        "points": trib_total,
        "meta": trib_total_meta,
        "chart_type": "line",
        "y_label": "Millones de $ (base 2001)",
        "intro": "Recaudación tributaria total mensual (AFIP, vía datos.gob.ar / Ministerio de Economía).",
    })
    indicators.append({
        "key": "iva",
        "title": "Ingresos tributarios — IVA",
        "etapas": [3],
        "points": iva,
        "meta": iva_meta,
        "chart_type": "line",
        "y_label": "Millones de $ (base 2001)",
        "intro": "Recaudación mensual del IVA.",
    })
    indicators.append({
        "key": "ganancias",
        "title": "Ingresos tributarios — Ganancias",
        "etapas": [3],
        "points": gan,
        "meta": gan_meta,
        "chart_type": "line",
        "y_label": "Millones de $ (base 2001)",
        "intro": "Recaudación mensual del Impuesto a las Ganancias.",
    })
    indicators.append({
        "key": "presion_tributaria",
        "title": "Presión tributaria (ingresos tributarios / PIB)",
        "etapas": [3],
        "points": presion,
        "meta": {
            "id": "calc_presion",
            "description": "Ingresos tributarios totales mensuales sobre PIB trimestral (interpolado mensualmente)",
            "source": "AFIP / INDEC (vía datos.gob.ar)",
            "units": "Porcentaje",
            "dataset": "142.3_TOTAL_2001_M_26 y 9.2_PPC_2004_T_22",
            "frequency": "Mensual (PIB trimestral expandido)",
        },
        "chart_type": "line",
        "y_label": "% del PIB",
        "intro": "Proxy de presión tributaria: recaudación tributaria total sobre PIB a precios corrientes.",
    })

    # 6. Tipo de cambio nominal — variación mensual (BCRA)
    tc_level, tc_meta = fetch_series("92.1_TCV_0_0_21")
    tc_var = mom_pct(tc_level)
    indicators.append({
        "key": "tc_nominal",
        "title": "Variación mensual del tipo de cambio nominal (pesos por dólar)",
        "etapas": [1, 2, 3],
        "points": tc_var,
        "meta": {**tc_meta, "note": "Variación % mensual calculada sobre promedio mensual del tipo de cambio de referencia (serie 92.1_TCV_0_0_21, BCRA)."},
        "chart_type": "line",
        "y_label": "% mensual",
        "intro": "Variación porcentual mensual del tipo de cambio nominal peso/dólar.",
    })

    # 7. Tipo de cambio real multilateral — BCRA ITCRM
    tcr, tcr_meta = fetch_series("116.4_TCRZE_2015_D_36_4", collapse="month")
    tcr_var = mom_pct(tcr)
    indicators.append({
        "key": "tc_real",
        "title": "Variación mensual del tipo de cambio real multilateral (ITCRM 2015=100)",
        "etapas": [1, 2, 3],
        "points": tcr_var,
        "meta": {**tcr_meta, "note": "Variación % mensual sobre ITCRM base 2015 (promedio mensual de serie diaria BCRA)."},
        "chart_type": "line",
        "y_label": "% mensual",
        "intro": "Apreciación/depreciación real multilateral del peso según el ITCRM del BCRA.",
    })

    # 8. Tasas de interés — BCRA
    badlar, badlar_meta = fetch_series("89.1_IR_BCRARIA_0_M_34")
    tib, tib_meta = fetch_series("89.1_TIB_0_0_20")
    indicators.append({
        "key": "badlar",
        "title": "Tasa de interés BADLAR bancos privados (nominal anual)",
        "etapas": [1, 3],
        "points": badlar,
        "meta": badlar_meta,
        "chart_type": "line",
        "y_label": "% nominal anual",
        "intro": "Tasa BADLAR de bancos privados en pesos (promedio mensual, BCRA).",
    })
    indicators.append({
        "key": "tib",
        "title": "Tasa de interés de depósitos a plazo fijo (TIB)",
        "etapas": [3],
        "points": tib,
        "meta": tib_meta,
        "chart_type": "line",
        "y_label": "% nominal anual",
        "intro": "Tasa de interés de depósitos a plazo fijo para el sector privado (BCRA).",
    })

    # 9. Gini — INDEC
    gini, gini_meta = fetch_series("65.1_CGI_0_0_21")
    gini_m = forward_fill_monthly(gini)
    indicators.append({
        "key": "gini",
        "title": "Desigualdad del ingreso — Coeficiente de Gini",
        "etapas": [1, 3],
        "points": gini_m,
        "meta": {**gini_meta, "note": "Serie trimestral EPH; valores replicados dentro de cada trimestre."},
        "chart_type": "line",
        "y_label": "Coeficiente de Gini",
        "intro": "Coeficiente de Gini del ingreso per cápita familiar (INDEC, EPH).",
    })

    # 10. Participación del ingreso laboral — INDEC Cuentas Nacionales
    rem_q, rem_meta = fetch_series("323.1_TOTAL_GENEADO__45")
    part_q = ratio_series(rem_q, pib_q)
    part_m = expand_quarterly_to_monthly(part_q)
    indicators.append({
        "key": "ingreso_laboral",
        "title": "Participación del ingreso laboral en el PIB",
        "etapas": [1, 3],
        "points": part_m,
        "meta": {
            "id": "323.1_TOTAL_GENEADO__45 / 9.2_PPC_2004_T_22",
            "description": "Remuneración al trabajo asalariado sobre PIB a precios corrientes",
            "source": "Instituto Nacional de Estadística y Censos (INDEC)",
            "units": "Porcentaje del PIB",
            "dataset": "Componentes del VAB y Producto Bruto Interno",
            "frequency": "Trimestral expandido mensualmente",
        },
        "chart_type": "line",
        "y_label": "% del PIB",
        "intro": "Participación de la remuneración al trabajo asalariado en el PIB (cuentas nacionales INDEC).",
    })

    # 11. Deuda pública / PIB — MECON
    deuda_pib, deuda_meta = read_deuda_pib_annual()
    indicators.append({
        "key": "deuda_pib",
        "title": "Deuda pública en relación al PIB",
        "etapas": [2, 3],
        "points": deuda_pib,
        "meta": deuda_meta,
        "chart_type": "line",
        "y_label": "% del PIB",
        "intro": "Deuda bruta de la Administración Central como porcentaje del PIB (Informes trimestrales de deuda, MECON).",
    })

    # Filter to 2023-01 .. latest 2026
    for ind in indicators:
        ind["points"] = [(d, round(v, 4) if isinstance(v, float) else v) for d, v in ind["points"] if d >= "2023-01" and d <= "2026-12"]
        if ind["points"]:
            vals = [v for _, v in ind["points"]]
            ind["stats"] = {
                "min": min(vals),
                "max": max(vals),
                "last": ind["points"][-1],
                "n": len(ind["points"]),
            }
            ind["conclusion"] = conclusion_text(ind)
        else:
            ind["stats"] = {"n": 0}
            ind["conclusion"] = "Sin datos oficiales disponibles en el período solicitado."

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "periodo": "Enero 2023 — último mes disponible 2026",
        "grupo": "A",
        "indicators": indicators,
    }


def conclusion_text(ind):
    pts = ind["points"]
    if len(pts) < 2:
        return "Serie insuficiente para evaluar tendencia."
    first, last = pts[0][1], pts[-1][1]
    delta = last - first
    direction = "al alza" if delta > 0 else "a la baja" if delta < 0 else "estable"
    return (
        f"Entre {pts[0][0]} y {pts[-1][0]} el indicador pasó de {fmt(first)} a {fmt(last)} "
        f"({direction}). Último dato oficial: {pts[-1][0]}."
    )


def fmt(v):
    if isinstance(v, float):
        if abs(v) >= 1e6:
            return f"{v:,.0f}"
        if abs(v) < 10:
            return f"{v:.2f}"
        return f"{v:,.2f}"
    return str(v)


def render_indicator_card(ind, suffix=""):
    sid = ind["key"] + suffix
    badges = "".join(f'<span class="badge">Etapa {e}</span>' for e in ind["etapas"])
    meta = ind["meta"]
    stats = ind.get("stats", {})
    last = stats.get("last")
    last_txt = f"{last[0]}: {fmt(last[1])}" if last else "—"
    rows = "".join(
        f"<tr><td>{p}</td><td>{fmt(v)}</td></tr>" for p, v in ind["points"]
    )
    note = meta.get("note", "")
    note_html = f'<br><em>Nota:</em> {note}' if note else ""
    url = meta.get("url", "")
    url_html = (
        f' URL: <a href="{url}" style="color:var(--accent)">{url}</a>' if url else ""
    )
    chart_id = f"chart-{sid}"
    return f"""
    <article class="card" id="{sid}" data-series-key="{ind['key']}">
      <h3>{ind['title']}</h3>
      {badges}
      <p class="intro">{ind['intro']}</p>
      <dl class="meta-grid">
        <div><dt>Fuente</dt><dd>{meta.get('source', '—')}</dd></div>
        <div><dt>Dataset</dt><dd>{meta.get('dataset', '—')}</dd></div>
        <div><dt>Unidades</dt><dd>{meta.get('units', '—')}</dd></div>
        <div><dt>Serie ID</dt><dd><code>{meta.get('id', '—')}</code></dd></div>
        <div><dt>Observaciones</dt><dd>{stats.get('n', 0)}</dd></div>
        <div><dt>Último valor</dt><dd>{last_txt}</dd></div>
      </dl>
      <div class="chart-wrap" id="{chart_id}" data-chart-type="{ind['chart_type']}" data-y-label="{ind['y_label']}"></div>
      <div class="conclusion"><strong>Conclusión:</strong> {ind.get('conclusion', '')}</div>
      <div class="table-scroll">
        <table>
          <thead><tr><th>Período</th><th>{ind['y_label']}</th></tr></thead>
          <tbody>{rows}</tbody>
        </table>
      </div>
      <div class="source">
        <strong>Fuente:</strong> {meta.get('source', '—')} — {meta.get('description', '')}.{url_html}{note_html}
      </div>
    </article>"""


def render_html(data):
    indicators = data["indicators"]
    indicators_json = json.dumps(indicators, ensure_ascii=False)

    etapa1 = [i for i in indicators if 1 in i["etapas"]]
    etapa2 = [i for i in indicators if 2 in i["etapas"]]
    etapa3 = [i for i in indicators if 3 in i["etapas"]]

    nav_links = (
        '<a href="#etapa1">Etapa 1</a>'
        '<a href="#etapa2">Etapa 2</a>'
        '<a href="#etapa3">Etapa 3</a>'
        '<a href="#etapa4">Etapa 4</a>'
        + "".join(f'<a href="#{i["key"]}">{i["key"]}</a>' for i in indicators)
    )

    consolidado_rows = "".join(
        f"""<tr>
          <td>{i['title']}</td>
          <td>{i['meta'].get('source', '—')}</td>
          <td>{i['stats']['last'][0] if i.get('stats', {}).get('last') else '—'}</td>
          <td>{fmt(i['stats']['last'][1]) if i.get('stats', {}).get('last') else '—'}</td>
          <td>{', '.join(str(e) for e in i['etapas'])}</td>
        </tr>"""
        for i in indicators
    )

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>STP Grupo A — Indicadores 2023-2026</title>
  <style>
    :root {{
      --bg: #0f1419;
      --card: #1a2332;
      --text: #e7ecf3;
      --muted: #8b9cb3;
      --accent: #3d8bfd;
      --border: #2a3548;
      --pos: #3dd68c;
      --neg: #f87171;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "IBM Plex Sans", "Segoe UI", system-ui, sans-serif;
      background: var(--bg);
      color: var(--text);
      line-height: 1.5;
    }}
    header {{
      padding: 2rem clamp(1rem, 4vw, 3rem);
      border-bottom: 1px solid var(--border);
      background: linear-gradient(135deg, #1a2332 0%, #0f1419 100%);
    }}
    h1 {{ margin: 0 0 .5rem; font-size: clamp(1.5rem, 3vw, 2.2rem); }}
    .subtitle {{ color: var(--muted); max-width: 70ch; }}
    nav {{
      position: sticky; top: 0; z-index: 10;
      background: rgba(15,20,25,.92); backdrop-filter: blur(8px);
      border-bottom: 1px solid var(--border);
      padding: .75rem clamp(1rem, 4vw, 3rem);
      display: flex; flex-wrap: wrap; gap: .5rem;
    }}
    nav a {{
      color: var(--accent); text-decoration: none; font-size: .85rem;
      padding: .25rem .6rem; border: 1px solid var(--border); border-radius: 999px;
    }}
    main {{ padding: 1.5rem clamp(1rem, 4vw, 3rem) 3rem; }}
    section {{ margin-bottom: 3rem; }}
    h2 {{ font-size: 1.35rem; margin: 2rem 0 1rem; color: #fff; }}
    h3 {{ font-size: 1.05rem; margin: 0 0 .75rem; }}
    .card {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 12px;
      padding: 1.25rem;
      margin-bottom: 1.5rem;
    }}
    .meta-grid {{
      display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
      gap: .75rem; margin: 1rem 0; font-size: .9rem;
    }}
    .meta-grid dt {{ color: var(--muted); font-size: .75rem; text-transform: uppercase; letter-spacing: .04em; }}
    .meta-grid dd {{ margin: .15rem 0 0; }}
    .chart-wrap {{ width: 100%; height: 320px; margin: 1rem 0; }}
    table {{
      width: 100%; border-collapse: collapse; font-size: .82rem;
      margin-top: 1rem;
    }}
    th, td {{ border: 1px solid var(--border); padding: .45rem .6rem; text-align: right; }}
    th {{ background: #243044; position: sticky; top: 0; }}
    td:first-child, th:first-child {{ text-align: left; }}
    .source {{
      margin-top: .75rem; padding-top: .75rem; border-top: 1px solid var(--border);
      font-size: .82rem; color: var(--muted);
    }}
    .source strong {{ color: var(--text); }}
    .badge {{
      display: inline-block; font-size: .7rem; padding: .15rem .5rem;
      border-radius: 4px; background: #243044; color: var(--accent); margin-right: .35rem;
    }}
    .intro {{ color: var(--muted); font-size: .9rem; margin-top: .5rem; }}
    .conclusion {{ background: #243044; border-radius: 8px; padding: .75rem 1rem; margin-top: 1rem; font-size: .92rem; }}
    .table-scroll {{ max-height: 280px; overflow: auto; border: 1px solid var(--border); border-radius: 8px; }}
    footer {{ padding: 2rem; color: var(--muted); font-size: .8rem; border-top: 1px solid var(--border); }}
  </style>
</head>
<body>
  <header>
    <h1>STP — Grupo A · Taller de Datos UT2-1</h1>
    <p class="subtitle">
      Valores mensuales de indicadores oficiales (INDEC, BCRA, Ministerio de Economía), período 2023–2026.
      Generado: {data["generated_at"]}. {data["periodo"]}.
    </p>
  </header>
  <main>
    <nav>{nav_links}</nav>

    <section id="etapa1">
      <h2>Etapa 1 — Indicadores Grupo A (INDEC, BCRA, MECON)</h2>
      <p class="subtitle">Tasa de inflación, resultados fiscales, bienes personales, ingresos tributarios, tipos de cambio, tasas de interés, Gini y participación del ingreso laboral.</p>
      {"".join(render_indicator_card(i, "") for i in etapa1)}
    </section>

    <section id="etapa2">
      <h2>Etapa 2 — Variables Grupo A</h2>
      <p class="subtitle">Resultado fiscal global y primario, tipos de cambio nominal y real, deuda pública/PIB.</p>
      {"".join(render_indicator_card(i, "-e2") for i in etapa2)}
    </section>

    <section id="etapa3">
      <h2>Etapa 3 — Informe por variable (gráfico + tabla + fuente + conclusión)</h2>
      {"".join(render_indicator_card(i, "-e3") for i in etapa3)}
    </section>

    <section id="etapa4">
      <h2>Etapa 4 — Consolidado Grupo A</h2>
      <div class="card">
        <table>
          <thead>
            <tr><th>Indicador</th><th>Fuente</th><th>Último período</th><th>Último valor</th><th>Etapa</th></tr>
          </thead>
          <tbody>{consolidado_rows}</tbody>
        </table>
      </div>
    </section>
  </main>
  <footer>
    Fuentes: INDEC, BCRA, Ministerio de Economía (IMIG, Informes de Deuda, AFIP vía datos.gob.ar).
    API: <a href="https://apis.datos.gob.ar/series/api/" style="color:var(--accent)">apis.datos.gob.ar/series</a>.
    Los huecos reflejan la última publicación oficial de cada serie.
  </footer>
  <script src="https://unpkg.com/react@18/umd/react.production.min.js"></script>
  <script src="https://unpkg.com/react-dom@18/umd/react-dom.production.min.js"></script>
  <script src="https://unpkg.com/recharts@2.12.7/umd/Recharts.js"></script>
  <script id="stp-data" type="application/json">{indicators_json}</script>
  <script>
    (function () {{
      const indicators = JSON.parse(document.getElementById('stp-data').textContent);

      function formatVal(v) {{
        if (v == null) return '—';
        if (Math.abs(v) >= 1e6) return v.toLocaleString('es-AR', {{ maximumFractionDigits: 0 }});
        return v.toLocaleString('es-AR', {{ maximumFractionDigits: 2 }});
      }}

      function mountCharts() {{
        if (!window.React || !window.ReactDOM || !window.Recharts) {{
          document.querySelectorAll('.chart-wrap').forEach(function (el) {{
            el.innerHTML = '<p style="color:#8b9cb3;padding:1rem">Gráfico no disponible (sin conexión a CDN). Los datos están en la tabla debajo.</p>';
          }});
          return;
        }}

        const R = window.React;
        const RD = window.ReactDOM;
        const RC = window.Recharts;
        const {{
          LineChart, Line, BarChart, Bar,
          XAxis, YAxis, CartesianGrid, Tooltip, Legend, ResponsiveContainer
        }} = RC;

        document.querySelectorAll('.chart-wrap[data-chart-type]').forEach(function (el) {{
          const seriesKey = el.parentElement.getAttribute('data-series-key');
          const ind = indicators.find(function (i) {{ return i.key === seriesKey; }});
          if (!ind || !ind.points || !ind.points.length) return;

          const chartData = ind.points.map(function (p) {{
            return {{ periodo: p[0], valor: p[1] }};
          }});

          const isBar = ind.chart_type === 'bar';
          const ChartCmp = isBar ? BarChart : LineChart;
          const SeriesCmp = isBar ? Bar : Line;

          const chart = R.createElement(ChartCmp, {{
            data: chartData,
            margin: {{ top: 10, right: 20, left: 10, bottom: 0 }}
          }},
            R.createElement(CartesianGrid, {{ strokeDasharray: '3 3', stroke: '#2a3548' }}),
            R.createElement(XAxis, {{
              dataKey: 'periodo',
              tick: {{ fill: '#8b9cb3', fontSize: 10 }},
              interval: 'preserveStartEnd',
              minTickGap: 30
            }}),
            R.createElement(YAxis, {{
              tick: {{ fill: '#8b9cb3', fontSize: 11 }},
              tickFormatter: formatVal,
              width: 80
            }}),
            R.createElement(Tooltip, {{
              contentStyle: {{ background: '#1a2332', border: '1px solid #2a3548' }},
              labelStyle: {{ color: '#e7ecf3' }},
              formatter: function (v) {{ return [formatVal(v), ind.y_label]; }}
            }}),
            R.createElement(Legend),
            R.createElement(SeriesCmp, {{
              type: 'monotone',
              dataKey: 'valor',
              name: ind.y_label,
              stroke: '#3d8bfd',
              fill: '#3d8bfd',
              strokeWidth: 2,
              dot: false
            }})
          );

          const tree = R.createElement(ResponsiveContainer, {{ width: '100%', height: 320 }}, chart);
          RD.createRoot(el).render(tree);
        }});
      }}

      if (document.readyState === 'loading') {{
        document.addEventListener('DOMContentLoaded', mountCharts);
      }} else {{
        mountCharts();
      }}
    }})();
  </script>
</body>
</html>
"""


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Build STP Grupo A outputs")
    parser.add_argument("--html", action="store_true", help="Generate HTML (default: all)")
    parser.add_argument("--json", action="store_true", help="Export JSON only")
    parser.add_argument("--notebook", action="store_true", help="Generate Jupyter notebook")
    parser.add_argument("--astro", action="store_true", help="Sync data for Astro site")
    args = parser.parse_args()
    all_outputs = not (args.html or args.json or args.notebook or args.astro)

    print("Fetching official data...")
    data = build_dataset()
    JSON_OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Written {JSON_OUT} ({JSON_OUT.stat().st_size // 1024} KB)")

    if all_outputs or args.html:
        html = render_html(data)
        OUT.write_text(html, encoding="utf-8")
        print(f"Written {OUT} ({OUT.stat().st_size // 1024} KB)")

    if all_outputs or args.notebook:
        from build_stp_notebook import write_notebook

        write_notebook(data)

    if all_outputs or args.astro:
        astro_data = ROOT / "stp-site" / "src" / "data" / "indicators.json"
        astro_data.parent.mkdir(parents=True, exist_ok=True)
        astro_data.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Written {astro_data}")

    for ind in data["indicators"]:
        print(f"  {ind['key']}: {ind['stats'].get('n', 0)} pts, last={ind['stats'].get('last')}")


if __name__ == "__main__":
    main()